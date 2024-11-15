import io
import openpyxl
from flask import send_file
from openpyxl.utils import get_column_letter
from flask_restful import Resource, reqparse, request
from models.cliente import ClienteModel
from models.usuario import UserModel
from flask_jwt_extended import jwt_required
from resources.filtros import normalize_path_params, consulta_com_codusuario_sem_status,consulta_com_codusuario_com_status,consulta_sem_codusuario_com_status,consulta_geral
from flask_jwt_extended import jwt_required
#import psycopg2
#from config import *
from datetime import datetime
import sqlite3

args = reqparse.RequestParser()
args.add_argument('nome', type=str, required=True,
                       help="The field 'nome' cannot be left blank")
args.add_argument('cpf', type=str, required=True,
                       help="The field 'cpf' cannot be left blank")
args.add_argument('datavis', type=str, required=True,
                       help="The field 'datavis' cannot be left blank")
args.add_argument('codqr', type=str, required=True,
                       help="The field 'codqr' cannot be left blank")
args.add_argument('codusuario', type=int, required=True,
                       help="The field 'codusuario' cannot be left blank")

class Clientes(Resource):
    @jwt_required()
    def get(self):
        check_dates = ClienteModel.checks_date()
        try:
            connection = sqlite3.connect(r"instance\banco.db") #psycopg2.connect(user=USER, password=PASSWORD, 
                         #                 host=HOST, port=PORT,
                         #                 database=DATABASE)
            cursor = connection.cursor()
        except sqlite3.Error as e:
            return {'message': f'Error accessing the database: {str(e)}'}, 500
        
        codusuario = request.args.get('codusuario', type=int)
        datamin = request.args.get('datamin', type=str)
        datamax = request.args.get('datamax', type=str)
        status = request.args.get('status', type=str)

        data = {
            'codusuario': codusuario,
            'datamin': datamin,
            'datamax': datamax,
            'status': status
        }
        valid_data = {key: value for key, value in data.items() if value}
        
        parameters = normalize_path_params(**valid_data)
        if not parameters.get('codusuario') and not parameters.get('status'):
            tupla = tuple([parameters[chave] for chave in parameters])
            cursor.execute(consulta_geral, tupla)
            result = cursor.fetchall()
        elif not parameters.get('status'):
            tupla = tuple([parameters[chave] for chave in parameters])
            cursor.execute(consulta_com_codusuario_sem_status, tupla)
            result = cursor.fetchall()
        elif not parameters.get('codusuario'):
            tupla = tuple([parameters[chave] for chave in parameters])
            cursor.execute(consulta_sem_codusuario_com_status, tupla)
            result = cursor.fetchall()
        else:
            tupla = tuple([parameters[chave] for chave in parameters])
            cursor.execute(consulta_com_codusuario_com_status, tupla)
            result = cursor.fetchall()

        clientes = []
        if result:
            for linha in result:
                #user = UserModel.find_user(linha[7]).json()
                #user = user["nome"]
                datager = datetime.strptime(linha[3], '%Y-%m-%d %H:%M:%S.%f').strftime('%d/%m/%Y %H:%M:%S')
                datavis = datetime.strptime(linha[4], '%Y-%m-%d %H:%M:%S.%f').strftime('%d/%m/%Y')
                dataaut = linha[6] 
                if dataaut is not None:
                    dataaut = datetime.strptime(dataaut, '%Y-%m-%d %H:%M:%S.%f').strftime('%d/%m/%Y %H:%M:%S')

                clientes.append({
                    'cliente_id': linha[0],
                    'nome': linha[1],
                    'cpf': linha[2],
                    'datager': datager,
                    'datavis': datavis,
                    'codqr': linha[5],
                    'dataaut': dataaut,
                    'codusuario': linha[7],
                    'status': linha[8],
                    #'usercad': user,
                })

        return {'clientes': clientes}


# Método para gerar e baixar arquivo Excel
    def generate_excel(self, clientes):
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Cabeçalhos do Excel
        headers = ['Cliente ID', 'Nome', 'CPF', 'Data Geração', 'Data Visita', 'Código QR', 'Data Autenticação', 'Cod Usuário', 'Status']
        ws.append(headers)

        # Preenchendo os dados no Excel
        for cliente in clientes:
            ws.append([
                cliente['cliente_id'],
                cliente['nome'],
                cliente['cpf'],
                cliente['datager'],
                cliente['datavis'],
                cliente['codqr'],
                cliente['dataaut'] if cliente['dataaut'] is not None else "N/A",
                cliente['codusuario'],
                cliente['status'],
                #cliente['usercad'],
            ])

        # Ajustando a largura das colunas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # Salvando o arquivo Excel em memória
        wb.save(output)
        output.seek(0)

        return output

    # Método do Flask-RESTful para o download
    def download_excel(self):
        # Utilizando os parâmetros da URL no download
        codusuario = request.args.get('codusuario', type=int)
        datamin = request.args.get('datamin', type=str)
        datamax = request.args.get('datamax', type=str)
        status = request.args.get('status', type=str)

        data = {
            'codusuario': codusuario,
            'datamin': datamin,
            'datamax': datamax,
            'status': status
        }
        valid_data = {key: value for key, value in data.items() if value}

        # Obter dados filtrados usando o método get()
        clientes_data = self.get()['clientes']
        
        if not clientes_data:
            return {"message": "No data available to download"}, 404

        # Gerar arquivo Excel com os dados
        output = self.generate_excel(clientes_data)

        # Enviar o arquivo Excel para download
        return send_file(output, as_attachment=True, download_name="clientes.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# Adicionando rota para download no Flask-RESTful
class DownloadClientes(Resource):
    @jwt_required()
    def get(self):
        clientes_resource = Clientes()
        return clientes_resource.download_excel()


class Cliente(Resource):
    @jwt_required()
    def post(self):
        data = args.parse_args()
        data['datavis'] = datetime.strptime(data['datavis'],"%Y-%m-%d") 
        if data['datavis'] < datetime.now():
                return {"message": F"Unable to register customer,date = {data['datavis'].strftime('%d/%m/%Y')} cannot be lower than the current date"},400
        if ClienteModel.find_cliente_by_cpf(data['cpf']):
            if ClienteModel.find_datavis(data['cpf'],datavis=data['datavis']):
                return {'message': f'There is already an appointment for the customer on this date: {data["datavis"].strftime("%d/%m/%Y")}'}, 400
        cliente = ClienteModel(**data, status="agendado")
        cliente.save_cliente()
        return cliente.json()

class Update_Cliente(Resource):
    args = reqparse.RequestParser()
    args.add_argument('datavis', type=str, required=True,
                       help="The field 'datavis' cannot be left blank")
    
    @jwt_required()
    def get(self, cliente_id):
        cliente = ClienteModel.find_cliente_by_id(cliente_id)
        if cliente:
            return cliente.json()
        return {'message': 'Cliente not found.'}, 404    
    
    @jwt_required()
    def put(self, cliente_id):
        data = Update_Cliente.args.parse_args()
        data['datavis'] = datetime.strptime(data['datavis'],"%Y-%m-%d")
        cliente_enc = ClienteModel.find_cliente_by_id(cliente_id)
        
        if cliente_enc:
            if ClienteModel.find_customer_status(cliente_id, "omisso"):
                return {'message': 'Unable to change the date, the customer did not show up'},400
            
            if ClienteModel.find_customer_status(cliente_id, "cancelado"):
                return {'message': 'Unable to change the date, the appointment is canceled'},400
        cliente_enc.update_cliente(**data)
        cliente_enc.save_cliente()
        return cliente_enc.json(), 200
        
"""     @jwt_required()
    def delete(self, cliente_id):
        cliente = ClienteModel.find_cliente(cliente_id)
        if cliente:
            cliente.delete_cliente()
            return {'message': 'Cliente deleted.'}
        return {'message': 'Cliente not found.'}, 404 """
    
class Cancels_Appointment(Resource):
    args = reqparse.RequestParser()
    args.add_argument('status', type=str, required=True,
                       help="The field 'status' cannot be left blank")   
    
    @jwt_required()
    def put(self, cliente_id):
        data = Cancels_Appointment.args.parse_args()
        cliente_enc = ClienteModel.find_cliente_by_id(cliente_id)
        
        if cliente_enc:
            if ClienteModel.find_customer_status(cliente_id, "omisso"):
                return {'message': 'Unable to change the date, the customer did not show up'},400
            if ClienteModel.find_customer_status(cliente_id, "cancelado"):
                return {'message': 'The appointment is already canceled'},400
        cliente_enc.update_appointment(**data)
        cliente_enc.save_cliente()
        return cliente_enc.json(), 200

class Status_Clients(Resource):  
     @jwt_required()
     def get(self):
        codusuario = request.args.get('codusuario', type=int)
        result = ClienteModel.checks_status(codusuario)

        clientes = []
        if result:
            for linha in result:
                clientes.append({
                    'nome': linha.nome,
                    'codqr': linha.codqr,
                })

        return {'clientes': clientes}

